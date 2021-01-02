import sys,os
import sys,os
# mypath =os.path.dirname(os.path.dirname(os.path.abspath(__file__))) #写入项目路径
# sys.path.append(mypath)
from TCJK.common.fengzhuangdingwei import Base
from TCJK.common.loginout import loginOut
from TCJK.pageses.allpage.pages import LoginPage
from appium import webdriver
from TCJK.pageses.allpage.pages import SetPage
import openpyxl
des = {
        "platformName": "Android",  # 手机是android还是ios
        "deviceName": "ca548277",  # 手机名称，通过adb devices获得
        "platformVersion": "10",  # android版本号
        "appPackage": "com.xingin.xhs",  # app包名
        "appActivity": "com.xingin.xhs.activity.SplashActivity",  # apk第一个启动页
        "noReset": True  # 不重置手机app
    }
class Login():
    def login1(self):
        driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", des)
        # openpyxl读取xlsx文件
        book1 = openpyxl.load_workbook('D://redBook//TCJK//data//logindata.xlsx')
        sh1 = book1.get_sheet_by_name('Sheet1')
        #获取最带行数
        max_rows = sh1.max_row
        #获取最大列数
        max_cells = sh1.max_column
        # 定义空字段
        dict_data = {}

        #把第一行标题设成字典的key
        for i in range(1,max_cells+1):
            dict_data.setdefault(sh1.cell(1,i).value)

        for row in range(2,max_rows+1):
            for cell in range(1,max_cells +1):
                dict_data[sh1.cell(1,cell).value] = sh1.cell(row,cell).value
            bae = Base(driver)
            print(LoginPage.其他号码登录)
            click_dw = bae.click(LoginPage.其他号码登录)
            click_dl = bae.click(LoginPage.手机密码登录)
            print(dict_data['phone'],dict_data['password'])
            send_phone = bae.send(LoginPage.请输入手机号码, dict_data['phone'])
            send_pwd = bae.send(LoginPage.请输入密码, dict_data['password'])

            try:
                click_dllbtn = bae.click(LoginPage.同意协议并登录)
                result = bae.find(LoginPage.登录遇到问题)
                if dict_data['except'] in result.text:
                    sh1.cell(row,max_cells,'PASS')
                    print('木木大模大样大模大样在')
                else:
                    sh1.cell(row, max_cells,'FAIL')
            except:
                print('ssssssssssssssssss')
                sh1.cell(row, max_cells,'PASS')

            finally:
                book1.save('D://redBook//TCJK//data//logindata.xlsx')
                bae.click(SetPage.我)
                bae.click(SetPage.设置按钮)
                bae.click(SetPage.登出账户)
                bae.click(SetPage.确定)







if __name__=="__main__":
    a = Login()
    a.login1()